import os
import time
import traceback

import openpyxl as openpyxl
from django.contrib import messages
from django.contrib.auth import logout, authenticate, login
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.hashers import make_password
from django.db import transaction
from django.db.models import Sum
from django.forms import model_to_dict
from django.http import JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.views.decorators.cache import never_cache

from appuser.models import AppUser
from constants.models import Constant
from contact.api.serializer import ContactsWeberializer
from contact.models import Contact
from mobile.models import Mobile
from payments.models import Transaction
from payments.utils import mpesa
from school.api.serializer import SchoolWebSerializer
from school.models import School
from student.models import Student
from tespython import *
from tokens.models import Token
from web.forms import EditStudentForm, AddStudentForm, EditParentForm, AddParentForm, AddSchoolForm, EditSchoolForm, \
    EditMobileForm, EditAgentForm, AddAgentForm, EditSettingsForm, ImportStudentsExcelForm, ImportParentExcelForm, \
    MinutesForm, DevicesForm


def getDetails(request, user):
    summarydictionary = {}

    # Access AppUser model fields
    app_user = AppUser.objects.get(id=user.id)
    fullname = app_user.fullname
    userid = app_user.id

    if app_user.isadmin:
        summarydictionary['color'] = "#0041C4"
        summarydictionary['istheadmin'] = True
    elif app_user.isagent:
        summarydictionary['color'] = "#F15A24"
        summarydictionary['istheadmin'] = False

    if 'globalschoolid' in request.session and request.session['globalschoolid'] != None and request.session['globalschoolid'] != "":
        try:
            school = School.objects.get(id=request.session['globalschoolid'])
        except:
            pass
        print("Here 2323")
    else:
        print(f"There {app_user}")
        print(f"There {model_to_dict(app_user)}")
        print(f"There {app_user.school}")
        school = app_user.school
        print(f"There {school}")

    try:

        summarydictionary['fullname'] = fullname
        summarydictionary['userid'] = str(userid)[:5].upper()

        schoolname = None
        schoolmobile = None
        unusedtokens = 0
        tokensconsumed = 0

        try:
            schooldevices =  Mobile.objects.filter(school__id=school.id)
            sum_standingtoken = schooldevices.aggregate(total_standingtoken=Sum('standingtoken'))['total_standingtoken']
            unusedtokens = sum_standingtoken
            sum_tokensconsumed = schooldevices.aggregate(total_tokensconsumed=Sum('tokensconsumed'))['total_tokensconsumed']
            tokensconsumed = sum_tokensconsumed
            schoolname = school.name
        except:
            pass

        print(f"Found school {school}")


        registeredStudents = Student.objects.filter(school=school)
        numberofRegisteredStudents = len(registeredStudents)
        parents = Contact.objects.filter(contactuser__school=school)

        numberofparents = len(parents)

        summarydictionary['school'] = school
        summarydictionary['schoolname'] = schoolname
        summarydictionary['schoolmobile'] = schoolmobile
        summarydictionary['unusedtokens'] = round(unusedtokens, 2)
        summarydictionary['usedtokens'] = round(tokensconsumed, 2)

        summarydictionary['registeredStudents'] = registeredStudents
        summarydictionary['numberofRegisteredStudents'] = numberofRegisteredStudents
        summarydictionary['parents'] = parents
        summarydictionary['numberofparents'] = numberofparents

        students = Student.objects.filter(school=school)
        summarydictionary['studentslist'] = students

        schooltransactions = Transaction.objects.filter(status="COMPLETE", student__school=school)
        print(f"Hey you used transactions len is {len(schooltransactions)}")

        totalschoolrevenue = schooltransactions.aggregate(Sum('amount'))['amount__sum']
        summarydictionary['totalschoolrevenue'] = totalschoolrevenue

        try:
            minutespertoken = Constant.objects.get(school__id=school.id).minutespertokenOrequivalentminutes
            print(f"Minutespertoken is ${min}")
            relativetalketime = unusedtokens * minutespertoken
            summarydictionary['usedtokensmins'] = round((tokensconsumed) * minutespertoken, 2)
            summarydictionary['minutespertoken'] = minutespertoken
            summarydictionary['relativetalketime'] = round(relativetalketime, 2)
        except:
            schooldevices = Mobile.objects.filter(school__id=school.id)
            sum_standingtoken = schooldevices.aggregate(total_standingtoken=Sum('standingtoken'))['total_standingtoken']
            unusedtokens = sum_standingtoken

            sum_tokensconsumed = schooldevices.aggregate(total_tokensconsumed=Sum('tokensconsumed'))['total_tokensconsumed']
            tokensconsumed = sum_tokensconsumed
            sum_minutesconsumed = schooldevices.aggregate(total_minutesconsumed=Sum('minutesconsumed'))['total_minutesconsumed']
            minutesconsumed = sum_minutesconsumed

            relativetalketime = unusedtokens * minutespertoken
            summarydictionary['usedtokensmins'] = tokensconsumed
            summarydictionary['relativetalketime'] = minutesconsumed

    except Exception as exception:
        traceback_str = traceback.format_exc()
        print(f"This is the error {traceback_str}")
        pass

    return summarydictionary


@never_cache
def agenthomepage(request, schoolid=None):
    if request.user.is_authenticated:
        print("User is authenticated")
        user = request.user
        if schoolid:
            request.session['globalschoolid'] = schoolid
            print("User has passed school id")

        summarydictionary = getDetails(request, user)
    else:
        return redirect('loginpage')
    response = render(request, "agentdashboard.html", {"summary": summarydictionary})
    return response


@never_cache
def studentshomepage(request):
    if request.user.is_authenticated:
        user = request.user
        summarydictionary = getDetails(request, user)
        try:
            students = Student.objects.filter(school=summarydictionary['school'])
            summarydictionary['studentslist'] = students
        except:
            pass
    else:
        return redirect('loginpage')
    response = render(request, "agenttables.html", {"summary": summarydictionary})
    return response


@never_cache
def deletestudent(request, studentid):
    if request.user.is_authenticated:
        student = Student.objects.get(id=studentid)
        student.delete()
        return redirect('studentshomepage')
    else:
        return redirect('loginpage')

    response = render(request, "agenttables.html", {"summary": monthsummaryDict})
    return response


@never_cache
def addStudent(request):
    if request.user.is_authenticated:
        user = request.user
    else:
        return redirect('loginpage')

    summarydictionary = getDetails(request, user)

    if request.method == 'POST':
        form = AddStudentForm(data=request.POST)
        if form.is_valid():
            with transaction.atomic():
                fullname = form.cleaned_data.get('fullname').strip()
                kcpeindexnumber = form.cleaned_data.get('kcpeindexnumber').strip()
                registrationnumber = form.cleaned_data.get('registrationnumber').strip()
                schoolid = int(form.cleaned_data.get('hidden_school').split('-')[0].strip())
                school = School.objects.get(id=schoolid)

                email = f"{registrationnumber}@gmail.com"
                username = f"{registrationnumber}@gmail.com"
                password = registrationnumber
                confirmpassword = registrationnumber

                # Create a new AppUser instance
                user = AppUser(
                    email=email,
                    username=username,
                    password=make_password(password),
                    confirmpassword=make_password(confirmpassword),
                    fullname=fullname,
                    isstudent=True,
                    isadmin=False,
                    isparent=False,
                    isagent=False,
                    school=school,
                )

                # Save the new AppUser instance
                try:
                    user.save()
                    newUser = user
                except Exception as exception:
                    messages.error(request, 'An error occurred while creating student: ' + str(exception))
                    return redirect('addStudent')

                # Create a new Student instance and link it to the new AppUser instance
                student = Student(
                    fullname=fullname,
                    kcpeindexnumber=kcpeindexnumber,
                    registrationnumber=registrationnumber,
                    school=school,
                    user=newUser,
                    password=password,
                    confirmpassword=confirmpassword,
                    username=username,
                    email=email
                )

                # Save the new Student instance
                try:
                    student.save()
                except Exception as exception:
                    messages.error(request, 'An error occurred while creating student: ' + str(exception))
                    return redirect('addStudent')
                # Add the new Student instance to any existing Contacts
        return redirect('studentshomepage')

    else:
        form = AddStudentForm(
            initial={
                'school': summarydictionary['school'],
            }
        )
        summarydictionary['form'] = form

    response = render(request, "addstudent.html", {"summary": summarydictionary})
    return response


def parentshomepage(request):
    if request.user.is_authenticated:
        user = request.user
    else:
        return redirect('loginpage')

    summarydictionary = getDetails(request, user)

    contactlist = Contact.objects.filter(contactuser__school=summarydictionary['school'])
    serializer = ContactsWeberializer(contactlist, many=True)
    serialized_data = serializer.data
    summarydictionary['parents'] = serialized_data

    # loop through the summary and replace each 'contacts' list with a list of Contact objects
    for item in summarydictionary['parents']:
        for student in item['students']:
            contacts = [get_object_or_404(Contact, pk=pk) for pk in student['contacts']]
            student['contacts'] = contacts

    print(serialized_data)

    response = render(request, "parenttables.html", {"summary": summarydictionary})
    return response


@never_cache
def editparent(request, parentid):
    if request.user.is_authenticated:
        user = request.user
    else:
        return redirect('loginpage')

    summarydictionary = getDetails(request, user)
    parent = Contact.objects.get(id=parentid)
    if request.method == 'POST':
        form = EditParentForm(data=request.POST)
        if form.is_valid():
            parent.name = form.cleaned_data.get('name').strip()
            parent.mobiletwo = transform_phone_number(form.cleaned_data.get('mobiletwo')).strip()
            parent.mobile = transform_phone_number(form.cleaned_data.get('mobile')).strip()
            try:
                parent.save()  # save the updated parent object
            except Exception as exception:
                return redirect('editparent')
            return redirect('parentshomepage')
        else:
            messages.error(request, 'Invalid entry')
    else:
        form = EditParentForm(
            initial={
                'name': parent.name,
                'mobile': parent.mobile,
                'mobiletwo': parent.mobiletwo
            }
        )
    summarydictionary['form'] = form
    response = render(request, "editparent.html", {"summary": summarydictionary})
    return response


@never_cache
def deleteparent(request, parentid):
    if request.user.is_authenticated:
        parent = Contact.objects.get(id=parentid)
        parent.delete()
        return redirect('parentshomepage')
    else:
        return redirect('loginpage')

    response = render(request, "agenttables.html", {"summary": monthsummaryDict})
    return response


@never_cache
def addparent(request):
    if request.user.is_authenticated:
        user = request.user
    else:
        return redirect('loginpage')

    summarydictionary = getDetails(request, user)

    if request.method == 'POST':
        form = AddParentForm(data=request.POST)
        if form.is_valid():
            with transaction.atomic():
                name = form.cleaned_data.get('name').strip()
                mobile = transform_phone_number(form.cleaned_data.get('mobile')).strip()
                mobiletwo = transform_phone_number(form.cleaned_data.get('mobiletwo')).strip()
                email = form.cleaned_data.get('email').strip()

                email = email

                # Create a new AppUser instance
                user = AppUser(
                    email=f"{mobile}@gmail.com",
                    username=email,
                    password=make_password(mobile),
                    confirmpassword=make_password(mobile),
                    fullname=name,
                    isstudent=True,
                    isadmin=False,
                    isparent=True,
                    isagent=False,
                    school=summarydictionary['school']
                )

                try:
                    user.save()
                    newUser = user
                except Exception as exception:
                    messages.error(request, 'An error occurred while saving parent: ' + str(exception))
                    return redirect('addparent')

                contact = Contact(
                    name=name,
                    mobile=mobile,
                    mobiletwo=mobiletwo,
                    email=email,
                    contactuser=newUser
                )

                # Save the new Student instance
                try:
                    contact.save()
                    messages.error(request, "Parent added successfully!")
                    return redirect('parentshomepage')
                except Exception as exception:
                    return redirect('addparent')

            # Add the new Student instance to any existing Contacts
        return redirect('parentshomepage')

    else:
        form = AddParentForm()
        summarydictionary['form'] = form

    response = render(request, "addparent.html", {"summary": summarydictionary})
    return response


def transactionshomepage(request):
    if request.user.is_authenticated:
        user = request.user
    else:
        return redirect('loginpage')

    summarydictionary = getDetails(request, user)

    schooltransactions = Transaction.objects.filter(student__school=summarydictionary['school'])
    summarydictionary['transactions'] = schooltransactions

    response = render(request, "transactiontables.html", {"summary": summarydictionary})
    return response


# BEGINNING OF ADMIN
@never_cache
def adminstudentshomepage(request):
    if request.user.is_authenticated:
        user = request.user

        summarydictionary = getDetails(request, user)
        try:
            students = Student.objects.filter(school=summarydictionary['school'])
            summarydictionary['studentslist'] = students

            print("HERE ARE STUDENTS")
            print(students)
        except:
            pass

    else:
        return redirect('loginpage')

    response = render(request, "agenttables.html", {"summary": summarydictionary})
    return response


@never_cache
def deletestudent(request, studentid):
    if request.user.is_authenticated:
        student = Student.objects.get(id=studentid)
        student.delete()
        return redirect('studentshomepage')
    else:
        return redirect('loginpage')

    response = render(request, "agenttables.html", {"summary": monthsummaryDict})
    return response


@never_cache
def editStudent(request, studentid):
    if request.user.is_authenticated:
        user = request.user
        summarydictionary = getDetails(request, user)

    else:
        return redirect('loginpage')

    student = Student.objects.get(id=studentid)
    contactlist = Contact.objects.filter(contactuser__school=summarydictionary['school'])
    serializer = ContactsWeberializer(contactlist, many=True)
    serialized_data = serializer.data
    summarydictionary['parents'] = serialized_data
    studentParents = student.contacts.all()
    summarydictionary['studentparents'] = studentParents


    if request.method == 'POST':
        form = EditStudentForm(data=request.POST)
        if form.is_valid():
            try:
                student.kcpeindexnumber = form.cleaned_data.get('kcpeindexnumber').strip()
                student.fullname = form.cleaned_data.get('fullname').strip()
            except:
                messages.error(request, "kcpeindexnumber is required")
                return redirect('editstudent', studentid=studentid)
            try:
                student.save()  # save the updated student object
            except Exception as exception:
                return redirect('editstudent', studentid=studentid)
            return redirect('studentshomepage')
        else:
            messages.error(request, 'Invalid entry')
    else:
        form = EditStudentForm(
            initial={
                'fullname': student.fullname,
                'kcpeindexnumber': student.kcpeindexnumber,
                'registrationnumber': student.registrationnumber,
            }
        )
    summarydictionary['form'] = form
    summarydictionary['studentid'] = studentid
    summarydictionary['contacts'] = student.contacts.all()
    print(f"Contacts wake ni {student.contacts.all()}")
    response = render(request, "editstudent.html", {"summary": summarydictionary})
    return response


@never_cache
def activateStudent(request, studentid):
    if request.user.is_authenticated:
        user = request.user

        summarydictionary = getDetails(request, user)
        try:
            students = Student.objects.filter(school=summarydictionary['school'])
            summarydictionary['studentslist'] = students
            print(students)
        except:
            pass

    else:
        return redirect('loginpage')

    school = summarydictionary['school']
    student = Student.objects.get(id=studentid)
    user = request.user
    app_user = AppUser.objects.get(id=user.id)
    app_user_mobile = app_user.phone
    activationFee = Constant.objects.get(school = school).activationamount
    gateway = mpesa.MpesaGateway()

    summarydictionary['activationfee'] = activationFee
    summarydictionary['student'] = student
    summarydictionary['studentname'] = student.fullname
    summarydictionary['mobile'] = app_user_mobile

    if request.method == 'POST':
        mobile = app_user_mobile
        therequest = request.POST
        thedictionary = therequest.dict()

        for value in thedictionary.items():
            thestring = value[1]
            start_index = thestring.find('254')
            end_index = start_index + 12
            mobile = transform_phone_number(thestring[start_index:end_index])
            print(mobile)

        print("Arrived here")
        print("Arrived here")
        print("Arrived here")
        print(mobile)

        timestamp = time.time()
        gateway.stk_push_request(activationFee, mobile, studentid, app_user, "REGISTRATION", timestamp)

        iscomplete = False
        start_time = time.time()
        while not iscomplete and time.time() - start_time < 60:
            status = Transaction.objects.filter(timestamp=timestamp).get().status
            print(f"Checking -- {status}")
            if status == "CANCELLED" or status == "FAILED":
                iscomplete = True
                return JsonResponse({'success': False})
            elif status == "COMPLETE":
                iscomplete = True
                return JsonResponse({'success': True})

        if not iscomplete:
            return JsonResponse({'success': False})

        # return redirect('studentshomepage')
        return JsonResponse({'success': True})

    response = render(request, "activatestudent.html", {"summary": summarydictionary})
    return response


@never_cache
def addStudent(request):
    if request.user.is_authenticated:
        user = request.user
    else:
        return redirect('loginpage')

    summarydictionary = getDetails(request, user)

    if request.method == 'POST':
        form = AddStudentForm(data=request.POST)
        if form.is_valid():
            with transaction.atomic():
                fullname = form.cleaned_data.get('fullname').strip()
                try:
                    kcpeindexnumber = form.cleaned_data.get('kcpeindexnumber').strip()
                except:
                    messages.error(request, "KCPE Index Number is required")
                    return redirect('addStudent')
                try:
                    registrationnumber = form.cleaned_data.get('registrationnumber').strip()
                except:
                    messages.error(request, "Registration Number is required")
                    return redirect('addStudent')
                try:
                    schoolid = int(form.cleaned_data.get('hidden_school').split('-')[0].strip())
                except:
                    messages.error(request, "School is required")
                    return redirect('addStudent')

                school = School.objects.get(id=schoolid)
                email = f"{registrationnumber}@gmail.com"
                username = f"{registrationnumber}@gmail.com"
                password = registrationnumber
                confirmpassword = registrationnumber

                # Create a new AppUser instance
                user = AppUser(
                    email=email,
                    username=username,
                    password=make_password(password),
                    confirmpassword=make_password(confirmpassword),
                    fullname=fullname,
                    isstudent=True,
                    isadmin=False,
                    isparent=False,
                    isagent=False,
                    school=school,
                )

                # Save the new AppUser instance
                try:
                    user.save()
                    newUser = user
                except Exception as exception:
                    messages.error(request, 'An error occurred while creating object: ' + str(exception))
                    return redirect('addStudent')

                # Create a new Student instance and link it to the new AppUser instance
                student = Student(
                    fullname=fullname,
                    kcpeindexnumber=kcpeindexnumber,
                    registrationnumber=registrationnumber,
                    school=school,
                    user=newUser,
                    password=password,
                    confirmpassword=confirmpassword,
                    username=username,
                    email=email
                )

                # Save the new Student instance
                try:
                    student.save()
                except Exception as exception:
                    messages.error(request, 'An error occurred while creating Student: ' + str(exception))
                    return redirect('addStudent')
                # Add the new Student instance to any existing Contacts
        return redirect('studentshomepage')

    else:
        try:
            form = AddStudentForm(
                initial={
                    'school': summarydictionary['school'],
                }
            )
            summarydictionary['form'] = form
        except:
            pass

    response = render(request, "addstudent.html", {"summary": summarydictionary})
    return response


@never_cache
def editparent(request, parentid):
    if request.user.is_authenticated:
        user = request.user
    else:
        return redirect('loginpage')

    summarydictionary = getDetails(request, user)
    parent = Contact.objects.get(id=parentid)
    if request.method == 'POST':
        form = EditParentForm(data=request.POST)
        if form.is_valid():
            parent.name = form.cleaned_data.get('name').strip()
            parent.mobiletwo = transform_phone_number(form.cleaned_data.get('mobiletwo'))
            parent.mobile = transform_phone_number(form.cleaned_data.get('mobile')).strip()
            try:
                parent.save()  # save the updated parent object
            except Exception as exception:
                return redirect('editparent')
            return redirect('parentshomepage')
        else:
            messages.error(request, 'Invalid entry')
    else:
        form = EditParentForm(
            initial={
                'name': parent.name,
                'mobile': parent.mobile,
                'mobiletwo': parent.mobiletwo
            }
        )
    summarydictionary['form'] = form
    response = render(request, "editparent.html", {"summary": summarydictionary})
    return response


@never_cache
def deleteparent(request, parentid):
    if request.user.is_authenticated:
        parent = Contact.objects.get(id=parentid)
        parent.delete()
        return redirect('parentshomepage')
    else:
        return redirect('loginpage')

    response = render(request, "agenttables.html", {"summary": monthsummaryDict})
    return response


@never_cache
def addparent(request):
    if request.user.is_authenticated:
        user = request.user
    else:
        return redirect('loginpage')

    summarydictionary = getDetails(request, user)

    if request.method == 'POST':
        form = AddParentForm(data=request.POST)
        if form.is_valid():
            with transaction.atomic():
                name = form.cleaned_data.get('name').strip()
                mobile = transform_phone_number(form.cleaned_data.get('mobile')).strip()
                mobiletwo = transform_phone_number(form.cleaned_data.get('mobiletwo'))
                email = form.cleaned_data.get('email').strip()

                email = email

                # Create a new AppUser instance
                user = AppUser(
                    email=email,
                    username=email,
                    password=make_password(email),
                    confirmpassword=make_password(email),
                    fullname=name,
                    isstudent=True,
                    isadmin=False,
                    isparent=True,
                    isagent=False,
                    school=summarydictionary['school']
                )

                try:
                    user.save()
                    newUser = user
                except Exception as exception:
                    messages.error(request, exception)
                    return redirect('addparent')

                contact = Contact(
                    name=name,
                    mobile=mobile,
                    mobiletwo=mobiletwo,
                    email=email,
                    contactuser=newUser
                )

                # Save the new Student instance
                try:
                    contact.save()
                except Exception as exception:
                    return redirect('addparent')
                # Add the new Student instance to any existing Contacts
        return redirect('parentshomepage')

    else:
        form = AddParentForm()
        summarydictionary['form'] = form

    response = render(request, "addparent.html", {"summary": summarydictionary})
    return response


@never_cache
def adminhomepage(request):
    if request.user.is_authenticated:
        user = request.user
    else:
        return redirect('loginpage')

    summarydictionary = getDetails(request, user)

    mobilestokens = Mobile.objects.aggregate(total_standingtoken=Sum('standingtoken'))
    mobileminutes = Mobile.objects.aggregate(total_standingtoken=Sum('standingminutes'))

    print(f"Mobile Tokens is {mobilestokens}")

    totalstandingtokens = mobilestokens['total_standingtoken']
    totalstandingminutes = mobileminutes['total_standingminutes']
    relativetalktime = totalstandingminutes

    # Get the current month's start and end dates
    current_month_start = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    next_month_start = (current_month_start + timezone.timedelta(days=31)).replace(day=1)
    thismonthtransactionslist = Transaction.objects.filter(status="COMPLETE", date_created__gte=current_month_start,
                                                           date_created__lt=next_month_start)
    thismonthtotalincome = thismonthtransactionslist.aggregate(Sum('amount'))['amount__sum']

    # Print the count of transactions
    # Get the start and end dates of the last month
    now = timezone.now()
    last_month_end = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    last_month_start = (last_month_end - timezone.timedelta(days=1)).replace(day=1)
    lastmonthtransactionslist = Transaction.objects.filter(status="COMPLETE", date_created__gte=last_month_start,
                                                           date_created__lt=last_month_end)
    lastmonthtotalincome = lastmonthtransactionslist.aggregate(Sum('amount'))['amount__sum']

    numberofusers = AppUser.objects.count()
    activedevices = Mobile.objects.filter(active=True).count()

    summarydictionary['totalstandingtokens'] = round(totalstandingtokens, 2)
    summarydictionary['totalrelativetalktime'] = round(relativetalktime, 2)
    summarydictionary['thismonthtransactionslist'] = thismonthtransactionslist
    summarydictionary['thismonthtotalincome'] = thismonthtotalincome
    summarydictionary['lastmonthtransactionslist'] = lastmonthtransactionslist
    summarydictionary['lastmonthtotalincome'] = lastmonthtotalincome
    summarydictionary['numberofusers'] = numberofusers
    summarydictionary['activedevices'] = activedevices

    response = render(request, "adminindex.html", {"summary": summarydictionary})
    return response


@never_cache
def adminschoolpage(request):
    if request.user.is_authenticated:
        user = request.user
    else:
        return redirect('loginpage')

    summarydictionary = getDetails(request, user)

    schoollist = School.objects.all()
    serializer = SchoolWebSerializer(schoollist, many=True)
    summarydictionary['schoollist'] = serializer.data

    response = render(request, "adminschooltable.html", {"summary": summarydictionary})
    return response


@never_cache
def editschool(request, schoolid):
    school = School.objects.get(id=schoolid)
    if request.user.is_authenticated:
        user = request.user
    else:
        return redirect('loginpage')
    if request.method == 'POST':
        form = EditSchoolForm(data=request.POST)
        if form.is_valid():
            school.name = form.cleaned_data.get('name').strip()
            school.email = form.cleaned_data.get('email').strip()
            school.location = form.cleaned_data.get('location').strip()
            mobile = form.cleaned_data.get('hidden_mobile')
            try:
                school.save()  # save the updated school object
            except Exception as exception:
                messages.error(request, exception)
                return redirect('editschool', schoolid)
            return redirect('adminschoolpage')
        else:
            messages.error(request, 'Invalid entry')
    else:
        form = EditSchoolForm(
            initial={
                'name': school.name,
                'email': school.email,
                'location': school.location,
                # 'hidden_mobile': school.mobile,  # assign mobile to hidden_mobile
            }
        )

    summarydictionary = getDetails(request, user)
    summarydictionary['form'] = form
    response = render(request, "editschool.html", {"summary": summarydictionary})
    return response


@never_cache
def deleteschool(request, schoolid):
    if request.user.is_authenticated:
        school = School.objects.get(id=schoolid)
        school.delete()
        return redirect('adminschoolpage')
    else:
        return redirect('loginpage')

    response = render(request, "adminschooltable.html", {"summary": monthsummaryDict})
    return response


@never_cache
def addschool(request):
    if request.user.is_authenticated:
        user = request.user
    else:
        return redirect('loginpage')

    summarydictionary = getDetails(request, user)

    if request.method == 'POST':
        form = AddSchoolForm(data=request.POST)
        if form.is_valid():
            with transaction.atomic():
                name = form.cleaned_data.get('name').strip()
                location = form.cleaned_data.get('location').strip()
                email = form.cleaned_data.get('email').strip()

                newschool = School(
                    name=name,
                    location=location,
                    email=email,
                )

                try:
                    newschool.save()
                    messages.success(request, "School added successfully!")
                    return redirect('adminschoolpage')

                except Exception as exception:
                    messages.error(request, f"{exception}")
                    return redirect('addschool')

        return redirect('adminschoolpage')

    else:
        form = AddSchoolForm()
        summarydictionary['form'] = form

    response = render(request, "addschool.html", {"summary": summarydictionary})
    return response


@never_cache
def viewschool(request, schoolid):
    print(f"SChool Id is {schoolid}")
    if request.user.is_authenticated:
        return redirect('agenthomepage', schoolid=schoolid)
    else:
        return redirect('loginpage')


@never_cache
def adminhomepage(request):
    if request.user.is_authenticated:
        user = request.user
    else:
        return redirect('loginpage')

    summarydictionary = getDetails(request, user)
    
    mobiles = Mobile.objects.aggregate(total_tokensconsumed=Sum('tokensconsumed'))
    tokensconsumed = mobiles['total_tokensconsumed']

    mobiles = Mobile.objects.aggregate(total_minutesconsumed=Sum('minutesconsumed'))
    minutesconsumed = mobiles['total_minutesconsumed']

    relativetalktime = minutesconsumed

    # Get the current month's start and end dates
    current_month_start = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    next_month_start = (current_month_start + timezone.timedelta(days=31)).replace(day=1)
    thismonthtransactionslist = Transaction.objects.filter(status="COMPLETE", date_created__gte=current_month_start,
                                                           date_created__lt=next_month_start)
    thismonthtotalincome = thismonthtransactionslist.aggregate(Sum('amount'))['amount__sum']

    # Print the count of transactions
    # Get the start and end dates of the last month
    now = timezone.now()
    last_month_end = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    last_month_start = (last_month_end - timezone.timedelta(days=1)).replace(day=1)
    lastmonthtransactionslist = Transaction.objects.filter(status="COMPLETE", date_created__gte=last_month_start,
                                                           date_created__lt=last_month_end)
    lastmonthtotalincome = lastmonthtransactionslist.aggregate(Sum('amount'))['amount__sum']

    numberofusers = AppUser.objects.count()
    activedevices = Mobile.objects.filter(active=True).count()

    try:
        summarydictionary['totalrelativetalktime'] = round(relativetalktime, 2)
        summarydictionary['thismonthtransactionslist'] = thismonthtransactionslist
        summarydictionary['thismonthtotalincome'] = thismonthtotalincome
        summarydictionary['lastmonthtransactionslist'] = lastmonthtransactionslist
        summarydictionary['lastmonthtotalincome'] = lastmonthtotalincome
        summarydictionary['numberofusers'] = numberofusers
        summarydictionary['activedevices'] = activedevices
        summarydictionary['totalstandingtokens'] = round(tokensconsumed, 2)
    except:
        pass

    response = render(request, "adminindex.html", {"summary": summarydictionary})
    return response


@never_cache
def admindevicepage(request):
    if request.user.is_authenticated:
        user = request.user
    else:
        return redirect('loginpage')

    summarydictionary = getDetails(request, user)
    mobilelist = Mobile.objects.all()
    summarydictionary['mobilelist'] = mobilelist

    response = render(request, "adminmobiletable.html", {"summary": summarydictionary})
    return response



@never_cache
def agentdevicepage(request):
    if request.user.is_authenticated:
        user = request.user
    else:
        return redirect('loginpage')

    summarydictionary = getDetails(request, user)
    mobilelist = Mobile.objects.all()
    summarydictionary['mobilelist'] = mobilelist

    response = render(request, "agentschoolmobiletable.html", {"summary": summarydictionary})
    return response


@never_cache
def editdevice(request, mobileid):
    device = Mobile.objects.get(id=mobileid)
    if request.user.is_authenticated:
        user = request.user
    else:
        return redirect('loginpage')
    if request.method == 'POST':
        form = EditMobileForm(data=request.POST)
        if form.is_valid():
            device.active = form.cleaned_data.get('active')
            try:
                device.save()
            except Exception as exception:
                return redirect('editdevice')
            return redirect('admindevicepage')
        else:
            messages.error(request, 'Invalid entry')
    else:
        form = EditMobileForm(
            initial={
                'active': device.active,
            }
        )

    summarydictionary = getDetails(request, user)
    summarydictionary['form'] = form
    response = render(request, "editdevice.html", {"summary": summarydictionary})
    return response


@never_cache
def admintokenspage(request):
    if request.user.is_authenticated:
        user = request.user
    else:
        return redirect('loginpage')

    try:
        tokenconsumedsum = Mobile.objects.aggregate(total_tokensconsumed=Sum('tokensconsumed'))
        tokenconsumedsum = tokenconsumedsum['total_tokensconsumed']

        minutesconsumed = Mobile.objects.aggregate(total_minutesconsumed=Sum('minutesconsumed'))
        minutesconsumed = minutesconsumed['total_minutesconsumed']

        summarydictionary = getDetails(request, user)
        devicelist = Mobile.objects.all()
        summarydictionary['devicelist'] = devicelist
        print(f"Hallo {tokenconsumedsum}")
        summarydictionary['tokenconsumedsum'] = round(tokenconsumedsum, 2)
        summarydictionary['relativetalktime'] = round(minutesconsumed, 2)
    except:
        pass

    response = render(request, "admintokenstable.html", {"summary": summarydictionary})
    return response


@never_cache
def adminagentspage(request):
    if request.user.is_authenticated:
        user = request.user
    else:
        return redirect('loginpage')

    agentlist = AppUser.objects.filter(isagent=True)
    summarydictionary = getDetails(request, user)
    summarydictionary['agentlist'] = agentlist

    response = render(request, "adminagentstable.html", {"summary": summarydictionary})
    return response


@never_cache
def editagent(request, agentid):
    agent = AppUser.objects.get(id=agentid)
    if request.user.is_authenticated:
        user = request.user
    else:
        return redirect('loginpage')
    if request.method == 'POST':
        form = EditAgentForm(data=request.POST)
        if form.is_valid():
            agent.fullname = form.cleaned_data.get('fullname').strip()
            agent.phone = transform_phone_number(form.cleaned_data.get('phone')).strip()
            agent.school = form.cleaned_data.get('school')
            try:
                agent.save()
            except Exception as exception:
                return redirect('editagent')

            return redirect('adminagentspage')
        else:
            messages.error(request, 'Invalid entry')
    else:
        form = EditAgentForm(
            initial={
                'fullname': agent.fullname,
                'phone': agent.phone,
                'school': agent.school,
            }
        )

    summarydictionary = getDetails(request, user)
    summarydictionary['form'] = form
    response = render(request, "editagent.html", {"summary": summarydictionary})
    return response


@never_cache
def deleteagent(request, agentid):
    if request.user.is_authenticated:
        agent = AppUser.objects.get(id=agentid)
        agent.delete()
        return redirect('adminagentspage')
    else:
        return redirect('loginpage')


@never_cache
def addagent(request):
    if request.user.is_authenticated:
        user = request.user
    else:
        return redirect('loginpage')

    summarydictionary = getDetails(request, user)

    if request.method == 'POST':
        form = AddAgentForm(data=request.POST)
        if form.is_valid():
            fullname = form.cleaned_data.get('fullname')
            if not fullname:
                return redirect('addagent')
            fullname = form.cleaned_data.get('fullname').strip()
            phone = transform_phone_number(form.cleaned_data.get('phone'))
            if not phone:
                return redirect('addagent')
            phone = transform_phone_number(form.cleaned_data.get('phone')).strip()

            school = form.cleaned_data.get('school')
            if not school:
                return redirect('addagent')
            email = f"{phone}@gmail.com"
            username = f"{phone}@gmail.com"
            password = phone
            confirmpassword = phone

            # Create a new AppUser instance
            user = AppUser(
                email=email,
                username=username,
                password=make_password(password),
                confirmpassword=make_password(confirmpassword),
                fullname=fullname,
                isstudent=False,
                isadmin=False,
                isparent=False,
                isagent=True,
                school=school,
                phone = phone
            )


            # Save the new AppUser instance
            try:
                user.save()
                newUser = user
            except Exception as exception:
                print("Here here here")
                messages.error(request, exception)
                return redirect('addagent')

        return redirect('adminagentspage')

    else:
        form = AddAgentForm()
        summarydictionary['form'] = form

    response = render(request, "addagent.html", {"summary": summarydictionary})
    return response


@never_cache
def logoutView(request):
    if request.user.is_authenticated:
        if 'globalschoolid' in globals():
            global globalschoolid
            del globalschoolid
        logout(request)
        return redirect('loginpage')


@never_cache
def loginhomepage(request):
    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            # Authenticate the user
            username = form.cleaned_data.get('username').strip()
            password = form.cleaned_data.get('password').strip()
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                appuser = AppUser.objects.get(id=user.id)
                print(appuser)
                global istheadmin

                if appuser.isadmin:
                    istheadmin = True
                    return redirect('adminhomepage')
                elif appuser.isagent:
                    istheadmin = False
                    return redirect('agenthomepageminusid')
                else:
                    print("User is neither admin nor agent")
        else:
            messages.error(request, 'Invalid username or password')
    else:
        form = AuthenticationForm()

    summarydictionary = {}
    summarydictionary['form'] = form
    response = render(request, "login.html", {"summary": summarydictionary})
    return response


@never_cache
def addtoken(request):
    if request.user.is_authenticated:
        user = request.user
        summarydictionary = getDetails(request, user)
    else:
        return redirect('loginpage')

    school = summarydictionary['school']
    constant = Constant.objects.get(school = school)
    minutespertoken = constant.minutespertokenOrequivalentminutes
    shillingspertoken = constant.shillingspertokenOrequivalentshillings

    if request.method == 'POST':
        token_value = float(request.POST.get('numbeoftokens'))
        if token_value <= 0:
            message = "Token value must be greater than 0."
            messages.error(request, message)
            return redirect('adminhomepage')

        else:
            equivalentshillings = shillingspertoken * token_value
            token = Token(
                school = school,
                tokenamount=token_value,
                equivalentshillings=equivalentshillings
            )
            try:
                token.save()
            except Exception as exception:
                return redirect('addtoken')
            return redirect('settingshomepage')

    response = redirect('adminhomepage')
    return response


@never_cache
def deletetoken(request, tokenid):
    if request.user.is_authenticated:
        getDetails(request, request.user)
        token = Token.objects.get(id=tokenid)
        token.delete()
        return redirect('settingshomepage')
    else:
        return redirect('loginpage')


@never_cache
def tokenlist(request, studentid):
    if request.user.is_authenticated:
        user = request.user
        summarydictionary = getDetails(request, user)
    else:
        return redirect('loginpage')

    school = summarydictionary['school']
    tokenlist = Token.objects.filter(school = school)
    minutespertoken = Constant.objects.get(school = school).minutespertokenOrequivalentminutes
    summarydictionary['tokenlist'] = []
    
    print(f"Token list is {tokenlist}")

    if tokenlist:
        for token in tokenlist:
            token.relativeminutes = token.tokenamount * minutespertoken
            summarydictionary['tokenlist'].append(token)

        print(tokenlist)
        studentname = Student.objects.get(id=studentid).fullname
        summarydictionary['studentid'] = studentid
        # summarydictionary['studentname'] = studentname
        summarydictionary['studentname'] = studentname

        response = render(request, "admintokenlisttable.html", {"summary": summarydictionary})
        return response
    else:
        messages.error(request, f"School settings have not been completed!")
        
        return redirect('studentshomepage')


@never_cache
def tokenbuy(request, studentid, amount):
    gateway = mpesa.MpesaGateway()

    if request.user.is_authenticated:
        user = request.user
        summarydictionary = getDetails(request, user)
    else:
        return redirect('loginpage')

    student = Student.objects.get(id=studentid)
    summarydictionary['studentname'] = student.fullname
    summarydictionary['studentid'] = student.id
    summarydictionary['amount'] = amount

    appuser = AppUser.objects.get(id=request.user.id)
    mobile = appuser.phone
    summarydictionary['mobile'] = mobile
    purpose = "TOKEN"

    if request.method == 'POST':
        print("It is a post request")
        therequest = request.POST
        thedictionary = therequest.dict()

        print(thedictionary)

        for value in thedictionary.items():
            thestring = value[1]
            print(thestring)
            start_index = thestring.find('254')
            end_index = start_index + 12
            mobile = transform_phone_number(thestring[start_index:end_index])

        print("Arrived here")
        print("Arrived here")
        print("Arrived here")
        print(mobile)

        timestamp = time.time()
        gateway.stk_push_request(amount, mobile, studentid, appuser, purpose, timestamp)

        iscomplete = False
        start_time = time.time()
        while not iscomplete and time.time() - start_time < 60:
            status = Transaction.objects.filter(timestamp=timestamp).get().status
            print(f"Checking -- {status}")
            if status == "CANCELLED" or status == "FAILED":
                iscomplete = True
                return JsonResponse({'success': False})
            elif status == "COMPLETE":
                iscomplete = True
                return JsonResponse({'success': True})

        if not iscomplete:
            return JsonResponse({'success': False})

        # return redirect('studentshomepage')
        return JsonResponse({'success': True})

    else:
        print("It is not post")
    response = render(request, "buytokenforstudent.html", {"summary": summarydictionary})
    return response


@never_cache
def importStudent(request):
    if request.user.is_authenticated:
        user = request.user
    else:
        return redirect('loginpage')

    summarydictionary = getDetails(request, user)

    if request.method == 'POST':
        print("at least am here 0")
        form = ImportStudentsExcelForm(request.POST, request.FILES)
        if form.is_valid():
            print("at least am here 1")

            student_fullname_table_name = None
            student_firstname_table_name = None
            student_middlename_table_name = None
            student_lastname_table_name = None
            student_admission_number_table_name = None
            student_kcpeindex_number_table_name = None
            parent_fullname_table_name = None
            parent_phone_number_table_name = None
            contact_person_one_fullname_table_name = None
            contact_person_one_mobile_table_name = None
            contact_person_two_fullname_table_name = None
            contact_person_two_mobile_table_name = None
            contact_person_three_fullname_table_name = None
            contact_person_three_mobile_table_name = None

            try:
                student_fullname_table_name = form.cleaned_data.get('student_fullname_table_name').strip()
            except:
                pass
            try:
                student_firstname_table_name = form.cleaned_data.get('student_firstname_table_name').strip()
            except:
                pass
            try:
                student_middlename_table_name = form.cleaned_data.get('student_middlename_table_name').strip()
            except:
                pass
            try:
                student_lastname_table_name = form.cleaned_data.get('student_lastname_table_name').strip()
            except:
                pass
            try:
                student_admission_number_table_name = form.cleaned_data.get('student_admission_number_table_name').strip()
            except:
                pass
            try:
                student_kcpeindex_number_table_name = form.cleaned_data.get('student_kcpeindex_number_table_name').strip()
            except:
                pass
            try:
                parent_fullname_table_name = form.cleaned_data.get('parent_fullname_table_name').strip()
            except:
                pass
            try:
                parent_phone_number_table_name = form.cleaned_data.get('parent_phone_number_table_name').strip()
            except:
                pass
            try:
                contact_person_one_fullname_table_name = form.cleaned_data.get('contact_person_one_fullname_table_name').strip()
            except:
                pass
            try:
                contact_person_one_mobile_table_name = form.cleaned_data.get('contact_person_one_mobile_table_name').strip()
            except:
                pass
            try:
                contact_person_two_fullname_table_name = form.cleaned_data.get( 'contact_person_two_fullname_table_name').strip()
            except:
                pass
            try:
                contact_person_two_mobile_table_name = form.cleaned_data.get('contact_person_two_mobile_table_name').strip()
            except:
                pass
            try:
                contact_person_three_fullname_table_name = form.cleaned_data.get('contact_person_three_fullname_table_name').strip()
            except:
                pass
            try:
                contact_person_three_mobile_table_name = form.cleaned_data.get('contact_person_three_mobile_table_name').strip()
            except:
                pass

            header_row_number = form.cleaned_data.get('header_row_number')
            excel_file = request.FILES['excel_file']

            # save the file to the assets folder
            file_path = 'static/assets/' + excel_file.name
            if os.path.exists(file_path):
                os.remove(file_path)
            with open(file_path, 'wb+') as destination:
                for chunk in excel_file.chunks():
                    destination.write(chunk)

            # process the file and save the data to the database
            wb = openpyxl.load_workbook(filename='static/assets/' + excel_file.name)
            ws = wb.active
            headers = []
            for cell in ws[header_row_number]:
                headers.append(cell.value)

            print(f"Found headers = {headers}")

            student_fullname_column_index = None
            student_firstname_column_index = None
            student_middlename_column_index = None
            student_lastname_column_index = None
            student_admission_number_column_index = None
            student_kcpeindex_number_column_index = None
            parent_fullname_column_index = None
            parent_phone_number_column_index = None
            contact_person_one_fullname_column_index = None
            contact_person_one_mobile_column_index = None
            contact_person_two_fullname_column_index = None
            contact_person_two_mobile_column_index = None
            contact_person_three_fullname_column_index = None
            contact_person_three_mobile_column_index = None

            if student_fullname_table_name == None or student_fullname_table_name not in headers:
                pass
            else:
                student_fullname_column_index = headers.index(student_fullname_table_name)
            if student_firstname_table_name == None or student_firstname_table_name not in headers:
                pass
            else:
                student_firstname_column_index = headers.index(student_firstname_table_name)
            if student_middlename_table_name == None or student_middlename_table_name not in headers:
                pass
            else:
                student_middlename_column_index = headers.index(student_middlename_table_name)
            if student_lastname_table_name == None or student_lastname_table_name not in headers:
                pass
            else:
                student_lastname_column_index = headers.index(student_lastname_table_name)
            if student_admission_number_table_name == None or student_admission_number_table_name not in headers:
                pass
            else:
                student_admission_number_column_index = headers.index(student_admission_number_table_name)
            if student_kcpeindex_number_table_name == None or student_kcpeindex_number_table_name not in headers:
                pass
            else:
                student_kcpeindex_number_column_index = headers.index(student_kcpeindex_number_table_name)
            if parent_fullname_table_name == None or parent_fullname_table_name not in headers:
                pass
            else:
                parent_fullname_column_index = headers.index(parent_fullname_table_name)
            if parent_phone_number_table_name == None or parent_phone_number_table_name not in headers:
                pass
            else:
                parent_phone_number_column_index = headers.index(parent_phone_number_table_name)
            if contact_person_one_fullname_table_name == None or contact_person_one_fullname_table_name not in headers:
                pass
            else:
                contact_person_one_fullname_column_index = headers.index(contact_person_one_fullname_table_name)
            if contact_person_one_mobile_table_name == None or contact_person_one_mobile_table_name not in headers:
                pass
            else:
                contact_person_one_mobile_column_index = headers.index(contact_person_one_mobile_table_name)
            if contact_person_two_fullname_table_name == None or contact_person_two_fullname_table_name not in headers:
                pass
            else:
                contact_person_two_fullname_column_index = headers.index(contact_person_two_fullname_table_name)
            if contact_person_two_mobile_table_name == None or contact_person_two_mobile_table_name not in headers:
                pass
            else:
                contact_person_three_fullname_column_index = headers.index(contact_person_three_fullname_table_name)
            if contact_person_three_fullname_table_name == None or contact_person_three_fullname_table_name not in headers:
                pass
            else:
                contact_person_three_mobile_column_index = headers.index(contact_person_three_mobile_table_name)
            if contact_person_three_mobile_table_name == None or contact_person_three_mobile_table_name not in headers:
                pass
            else:
                contact_person_two_mobile_column_index = headers.index(contact_person_two_mobile_table_name)

            student_fullname = None
            student_firstname = None
            student_middlename = None
            student_lastname = None
            student_admission_number = None
            student_kcpeindex_number = None
            parent_fullname = None
            parent_phone_number = None
            contact_person_one_fullname = None
            contact_person_one_mobile = None
            contact_person_two_fullname = None
            contact_person_two_mobile = None
            contact_person_three_fullname = None
            contact_person_three_mobile = None

            for row in ws.iter_rows(min_row=(header_row_number + 1)):
                if student_fullname_column_index is not None:
                    student_fullname = row[student_fullname_column_index].value
                    print(f"OKAY SO WA KWANZA NI {student_fullname}")
                if student_firstname_column_index is not None:
                    student_firstname = row[student_firstname_column_index].value
                if student_middlename_column_index is not None:
                    student_middlename = row[student_middlename_column_index].value
                if student_lastname_column_index is not None:
                    student_lastname = row[student_lastname_column_index].value
                if student_admission_number_column_index is not None:
                    student_admission_number = row[student_admission_number_column_index].value
                if student_kcpeindex_number_column_index is not None:
                    student_kcpeindex_number = row[student_kcpeindex_number_column_index].value
                if parent_fullname_column_index is not None:
                    parent_fullname = row[parent_fullname_column_index].value
                if parent_phone_number_column_index is not None:
                    parent_phone_number = row[parent_phone_number_column_index].value
                if contact_person_one_fullname_column_index is not None:
                    contact_person_one_fullname = row[contact_person_one_fullname_column_index].value
                if contact_person_one_mobile_column_index is not None:
                    contact_person_one_mobile = transform_phone_number(
                        row[contact_person_one_mobile_column_index].value)
                if contact_person_two_fullname_column_index is not None:
                    contact_person_two_fullname = row[contact_person_two_fullname_column_index].value
                if contact_person_two_mobile_column_index is not None:
                    contact_person_two_mobile = transform_phone_number(
                        row[contact_person_two_mobile_column_index].value)
                if contact_person_three_fullname_column_index is not None:
                    contact_person_three_fullname = row[contact_person_three_fullname_column_index].value
                if contact_person_three_mobile_column_index is not None:
                    contact_person_three_mobile = transform_phone_number(
                        row[contact_person_three_mobile_column_index].value)

                newStudent = Student()

                if student_fullname:
                    newStudent.fullname = student_fullname
                elif student_firstname:
                    newStudent.fullname = student_firstname
                    if student_middlename:
                        newStudent.fullname = f"{student_firstname} {student_middlename}"
                        if student_lastname:
                            newStudent.fullname = f"{student_firstname} {student_middlename} {student_lastname}"
                elif student_middlename:
                    newStudent.fullname = student_middlename
                    if student_middlename:
                        newStudent.fullname = f"{student_firstname} {student_middlename}"
                        if student_lastname:
                            newStudent.fullname = f"{student_firstname} {student_middlename} {student_lastname}"
                elif student_lastname:
                    newStudent.fullname = student_lastname
                    if student_middlename:
                        newStudent.fullname = f"{student_firstname} {student_middlename}"
                        if student_lastname:
                            newStudent.fullname = f"{student_firstname} {student_middlename} {student_lastname}"
                else:
                    pass

                if student_admission_number:
                    newStudent.registrationnumber = student_admission_number
                    newStudent.email = f"{student_admission_number}@gmail.com"
                    newStudent.username = f"{student_admission_number}@gmail.com"
                    newStudent.email = f"{student_admission_number}@gmail.com"
                    newStudent.password = f"{student_admission_number}"
                    newStudent.confirmpassword = f"{student_admission_number}"

                    studentntUser = AppUser(
                        email=newStudent.email,
                        username=newStudent.username,
                        password=make_password(str(newStudent.password)),
                        confirmpassword=make_password(str(newStudent.confirmpassword)),
                        fullname=newStudent.username,
                        isstudent=True,
                        isadmin=False,
                        isparent=False,
                        isagent=False,
                        school=summarydictionary['school'],
                    )
                    try:
                        studentntUser = AppUser.objects.get(email=newStudent.email)
                    except AppUser.DoesNotExist:
                        studentntUser.save()
                        newStudent.user = studentntUser

                else:
                    pass

                if student_kcpeindex_number:
                    newStudent.kcpeindexnumber = student_kcpeindex_number

                contactuserparent = Contact()

                if parent_fullname:
                    contactuserparent.name = parent_fullname
                    if parent_phone_number:
                        contactuserparent.mobile = parent_phone_number
                        contactuserparent.relationship = "PARENT"
                        contactuserparent.email = f"{contactuserparent.mobile}@gmail.com"

                        parentuser = AppUser(
                            email=contactuserparent.email,
                            username=contactuserparent.email,
                            password=make_password(str(contactuserparent.mobile)),
                            confirmpassword=make_password(str(contactuserparent.mobile)),
                            fullname=contactuserparent.name,
                            isstudent=False,
                            isadmin=False,
                            isparent=True,
                            isagent=False,
                            school=summarydictionary['school'],
                        )

                        try:
                            parentuser = AppUser.objects.get(email=contactuserparent.email)
                        except AppUser.DoesNotExist:
                            parentuser.save()
                            contactuserparent.contactuser = parentuser
                            contactuserparent.save()
                    else:
                        pass

                contactuserone = Contact()
                if contact_person_one_fullname:
                    contactuserone.name = contact_person_one_fullname
                    if contact_person_one_mobile:
                        contactuserone.mobile = contact_person_one_mobile
                        contactuserone.relationship = "GUARDIAN"
                        contactuserone.email = f"{contactuserone.mobile}@gmail.com"
                        guardianuser = AppUser(
                            email=contactuserone.email,
                            username=contactuserone.email,
                            password=make_password(str(contactuserone.mobile)),
                            confirmpassword=make_password(str(contactuserone.mobile)),
                            fullname=contactuserone.name,
                            isstudent=False,
                            isadmin=False,
                            isparent=True,
                            isagent=False,
                            school=summarydictionary['school'])
                        try:
                            guardianuser = AppUser.objects.get(email=contactuserone.email)
                        except AppUser.DoesNotExist:
                            guardianuser.save()
                            contactuserone.contactuser = guardianuser
                            contactuserone.save()
                    else:
                        pass

                contactuserTwo = Contact()
                if contact_person_two_fullname:
                    contactuserTwo.name = contact_person_two_fullname
                    if contact_person_two_mobile:
                        contactuserTwo.mobile = contact_person_two_mobile
                        contactuserTwo.relationship = "GUARDIAN"
                        contactuserTwo.email = f"{contactuserTwo.mobile}@gmail.com"
                        guardianuser = AppUser(
                            email=contactuserTwo.email,
                            username=contactuserTwo.email,
                            password=make_password(str(contactuserTwo.mobile)),
                            confirmpassword=make_password(str(contactuserTwo.mobile)),
                            fullname=contactuserTwo.name,
                            isstudent=False,
                            isadmin=False,
                            isparent=True,
                            isagent=False,
                            school=summarydictionary['school'],
                        )
                        try:
                            guardianuser = AppUser.objects.get(email=contactuserTwo.email)
                        except AppUser.DoesNotExist:
                            guardianuser.save()
                            contactuserTwo.contactuser = guardianuser
                            contactuserTwo.save()

                    else:
                        pass

                contactuserThree = Contact()
                if contact_person_three_fullname:
                    contactuserThree.name = contact_person_three_fullname
                    if contact_person_three_mobile:
                        contactuserThree.mobile = contact_person_three_mobile
                        contactuserThree.relationship = "GUARDIAN"
                        contactuserThree.email = f"{contactuserThree.mobile}@gmail.com"
                        guardianuser = AppUser(
                            email=contactuserThree.email,
                            username=contactuserThree.email,
                            password=make_password(str(contactuserThree.mobile)),
                            confirmpassword=make_password(str(contactuserThree.mobile)),
                            fullname=contactuserThree.name,
                            isstudent=False,
                            isadmin=False,
                            isparent=True,
                            isagent=False,
                            school=summarydictionary['school'],
                        )
                        try:
                            guardianuser = AppUser.objects.get(email=contactuserThree.email)
                        except AppUser.DoesNotExist:
                            guardianuser.save()
                            contactuserThree.contactuser = guardianuser
                            contactuserThree.save()
                    else:
                        pass

                try:
                    newStudent.school = summarydictionary['school']
                    newStudent.save()
                    if hasattr(contactuserparent, 'id'):
                        newStudent.contacts.add(contactuserparent)
                    if hasattr(contactuserone, 'id'):
                        newStudent.contacts.add(contactuserone)
                    if hasattr(contactuserTwo, 'id'):
                        newStudent.contacts.add(contactuserTwo)
                    if hasattr(contactuserThree, 'id'):
                        newStudent.contacts.add(contactuserThree)
                except:
                    pass

            return JsonResponse({'success': True})

    else:
        try:
            form = ImportStudentsExcelForm()
            summarydictionary['form'] = form
        except:
            pass

    response = render(request, "agent_import_students.html", {"summary": summarydictionary})
    return response


@never_cache
def importParent(request):
    if request.user.is_authenticated:
        user = request.user
    else:
        return redirect('loginpage')

    summarydictionary = getDetails(request, user)

    if request.method == 'POST':
        form = ImportParentExcelForm(request.POST, request.FILES)
        if form.is_valid():

            parent_fullname_table_name = None
            parent_phone_number_table_name = None
            mobiletwo_table_name = None
            header_row_number = None

            try:
                parent_fullname_table_name = form.cleaned_data.get('parent_fullname_table_name').strip()
            except:
                pass
            try:
                parent_phone_number_table_name = form.cleaned_data.get('parent_phone_number_table_name').strip()
            except:
                pass
            try:
                mobiletwo_table_name = form.cleaned_data.get('mobiletwo_table_name').strip()
            except:
                pass
            try:
                header_row_number = form.cleaned_data.get('header_row_number')
            except:
                pass
            excel_file = request.FILES['excel_file']

            # save the file to the assets folder
            file_path = 'static/assets/' + excel_file.name
            if os.path.exists(file_path):
                os.remove(file_path)
            with open(file_path, 'wb+') as destination:
                for chunk in excel_file.chunks():
                    destination.write(chunk)

            # process the file and save the data to the database
            wb = openpyxl.load_workbook(filename='static/assets/' + excel_file.name)
            ws = wb.active
            headers = []
            for cell in ws[header_row_number]:
                headers.append(cell.value)

            print(f"Found headers = {headers}")

            parent_fullname_column_index = None
            parent_phone_number_column_index = None
            mobiletwo_column_index = None
            email_column_index = None

            if parent_fullname_table_name == None or parent_fullname_table_name not in headers:
                pass
            else:
                parent_fullname_column_index = headers.index(parent_fullname_table_name)
            if parent_phone_number_table_name == None or parent_phone_number_table_name not in headers:
                pass
            else:
                parent_phone_number_column_index = headers.index(parent_phone_number_table_name)
            if mobiletwo_table_name == None or mobiletwo_table_name not in headers:
                pass
            else:
                mobiletwo_column_index = headers.index(mobiletwo_table_name)
            if email_column_index == None or email_column_index not in headers:
                pass
            else:
                email_column_index = headers.index(email_column_index)

            parent_fullname = None
            parent_phone_number = None
            mobiletwo = None
            email = None

            for row in ws.iter_rows(min_row=(header_row_number + 1)):
                if parent_fullname_column_index is not None:
                    parent_fullname = row[parent_fullname_column_index].value
                if parent_phone_number_column_index is not None:
                    parent_phone_number = transform_phone_number(row[parent_phone_number_column_index].value)
                if mobiletwo_column_index is not None:
                    mobiletwo = transform_phone_number(row[mobiletwo_column_index].value)
                if email_column_index is not None:
                    email = row[email_column_index].value

                newParent = Contact()

                if parent_phone_number:
                    newParent.mobile = parent_phone_number
                    newParent.relationship = "PARENT"

                    if parent_fullname:
                        newParent.name = parent_fullname

                    if mobiletwo:
                        newParent.mobiletwo = mobiletwo

                    if email:
                        newParent.email = email

                    parentUser = AppUser(
                        email=f"{newParent.mobile}@gmail.com",
                        username=f"{newParent.mobile}@gmail.com",
                        password=make_password(str(newParent.mobile)),
                        confirmpassword=make_password(str(newParent.mobile)),
                        fullname=newParent.name,
                        isstudent=False,
                        isadmin=False,
                        isparent=True,
                        isagent=False,
                        school=summarydictionary['school'],
                    )

                    try:
                        parentUser = AppUser.objects.get(email=parentUser.email)
                    except AppUser.DoesNotExist:
                        parentUser.save()
                        newParent.contactuser = parentUser

                try:
                    newParent.save()
                except:
                    pass
                else:
                    pass

            return JsonResponse({'success': True})

    else:
        try:
            form = ImportParentExcelForm()
            summarydictionary['form'] = form
        except:
            pass

    response = render(request, "agent_import_parents.html", {"summary": summarydictionary})
    return response



@never_cache
def adddevice(request):
    if request.user.is_authenticated:
        user = request.user
    else:
        return redirect('loginpage')

    summarydictionary = getDetails(request, user)

    if request.method == 'POST':
        form = DevicesForm(data=request.POST)
        if form.is_valid():
            mobile = transform_phone_number(form.cleaned_data.get('mobile')).strip()
            school = form.cleaned_data.get('school')
            theschool = School.objects.get(id = school.id)

            print(f"----- {theschool}")
            if not mobile:
                messages.error(request, "You did not enter the mobile")
                return redirect('adddevice')

            try:
                Mobile.objects.get(mobile = mobile)
                messages.error(request, f"Mobile {mobile} already exists!")
                return redirect('adddevice')
            except Mobile.DoesNotExist:
                try:
                    newMobileInstance = Mobile.objects.create(mobile=mobile, school = theschool)
                    newMobileInstance.save()
                    messages.success(request, "Mobile saved successfully!")
                    return redirect('admindevicepage')
                except Exception as exception:
                    messages.error(request, exception)
                    return redirect('adddevice')

        else:
            print("Form is not valid")
            messages.error(request, "Form is not valid")
            return redirect('adddevice')


    else:
        form = DevicesForm()
        summarydictionary['form'] = form

    response = render(request, "adddevice.html", {"summary": summarydictionary})
    return response



@never_cache
def deleteDevice(request, mobileid):
    if request.user.is_authenticated:
        try:
            mobile = Mobile.objects.get(id=mobileid)
            mobile.delete()
            messages.success(request, f"Mobile deleted successfully")
            return redirect('admindevicepage')
        except Mobile.DoesNotExist:
            messages.error(request, f"Does Not Exist")
            return redirect('admindevicepage')
    else:
        return redirect('loginpage')



@never_cache
def schooldevices(request, schoolid):
    if request.user.is_authenticated:
        user = request.user
    else:
        return redirect('loginpage')

    summarydictionary = getDetails(request, user)
    mobilelist = Mobile.objects.filter(school_id=schoolid)
    summarydictionary['mobilelist'] = mobilelist

    print(f"Mobile List {mobilelist}")

    response = render(request, "schoolmobiletable.html", {"summary": summarydictionary})
    return response

@never_cache
def agentschooldevices(request, schoolid):
    if request.user.is_authenticated:
        user = request.user
    else:
        return redirect('loginpage')

    summarydictionary = getDetails(request, user)
    mobilelist = Mobile.objects.filter(school_id=schoolid)
    summarydictionary['mobilelist'] = mobilelist

    print(f"Mobile List {mobilelist}")

    response = render(request, "agentschoolmobiletable.html", {"summary": summarydictionary})
    return response

@never_cache
def addParentToStudent(request, studentid, parentid):
    if request.user.is_authenticated:
        user = request.user
    else:
        return redirect('loginpage')

    getDetails(request, user)
    try:
        parent = Contact.objects.get(id=parentid)
        student = Student.objects.get(id=studentid)
        student.contacts.add(parent)
        student.save()
    except (Contact.DoesNotExist, Student.DoesNotExist):
        parent = None
        student = None
    except Exception as e:
        pass

    return redirect('editstudent', studentid)


@never_cache
def removeParentFromStudent(request, studentid, parentid):
    if request.user.is_authenticated:
        user = request.user
    else:
        return redirect('loginpage')

    getDetails(request, user)
    try:
        parent = Contact.objects.get(id=parentid)
        student = Student.objects.get(id=studentid)
        student.contacts.remove(parent)
        student.save()
    except (Contact.DoesNotExist, Student.DoesNotExist):
        parent = None
        student = None
    except Exception as e:
        pass

    return redirect('editstudent', studentid)




@never_cache
def settingshomepage(request):
    if request.user.is_authenticated:
        user = request.user
        summarydictionary = getDetails(request, user)
    else:
        return redirect('loginpage')
    school = summarydictionary['school']

    setting = None
    try:
        setting = Constant.objects.get(school = school)
    except Constant.DoesNotExist:
        print("Settig Does Not Exist!")
        setting = Constant.objects.create(school=school)
        print(f"setting is {setting}")

    if request.method == 'POST':
        form = EditSettingsForm(data=request.POST)
        if form.is_valid():
            setting.activationamount = form.cleaned_data.get('activationamount')
            setting.minutespertokenOrequivalentminutes = form.cleaned_data.get('minutespertokenOrequivalentminutes')
            setting.minutepershilling = form.cleaned_data.get('minutepershilling')
            shillingspertokenOrequivalentshillings = round(setting.minutespertokenOrequivalentminutes / setting.minutepershilling, 2)

            setting.minutepershilling = form.cleaned_data.get('minutepershilling')
            setting.shillingspertokenOrequivalentshillings = shillingspertokenOrequivalentshillings
            try:
                print(f"School is {school}")
                print(f"Setting is {setting}")
                setting.save()
            except Exception as exception:
                print(f"Exception is {exception}")
                messages.error(request, exception)
                return redirect('settingshomepage')
            return redirect('settingshomepage')
        else:
            messages.error(request, 'Invalid entry')
    else:
        form = EditSettingsForm(
            initial={
                'activationamount': setting.activationamount,
                'minutespertokenOrequivalentminutes': setting.minutespertokenOrequivalentminutes,
                'minutepershilling': setting.minutepershilling,
            }
        )

    summarydictionary['activationamount'] = setting.activationamount
    summarydictionary['minutepershilling'] = setting.minutepershilling
    summarydictionary['minutespertoken'] = setting.minutespertokenOrequivalentminutes
    summarydictionary['form'] = form
    summarydictionary['shillingspertokenOrequivalentshillings'] = setting.shillingspertokenOrequivalentshillings

    tokenlist = Token.objects.filter(school = school)
    summarydictionary['tokenlist'] = tokenlist

    school = summarydictionary['school']
    schoolid = school.id
    mobilelist = Mobile.objects.filter(school_id=schoolid)
    summarydictionary['mobilelist'] = mobilelist

    response = render(request, "editsettings.html", {"summary": summarydictionary})
    return response



@never_cache
def addMinutesToDevice(request):
    if request.user.is_authenticated:
        user = request.user
        summarydictionary = getDetails(request, user)
    else:
        return redirect('loginpage')
    school = summarydictionary['school']

    if request.method == 'POST':
        mobilenumber = request.POST.get('mobileid')
        formminutes = request.POST.get('minutes')

        if formminutes and mobilenumber:
            minutes = float(formminutes)

            try:
                mobileInstance = Mobile.objects.get(mobile=mobilenumber)

                try:
                    school = mobileInstance.school
                    try:
                        try:
                            minutespertoken = Constant.objects.get(school=school).minutespertokenOrequivalentminutes
                        except:
                            messages.error(request, f"Please add settings for {school}")
                            return redirect('settingshomepage')
                        else:
                            tokens = minutes / minutespertoken if minutespertoken else 0
                            oldstandingtoken = mobileInstance.standingtoken
                            oldstandingminutes = mobileInstance.standingminutes

                            newminutes = oldstandingminutes + minutes
                            newtokens = oldstandingtoken + tokens

                            mobileInstance.standingtoken = newtokens
                            mobileInstance.standingminutes = newminutes
                    except Constant.DoesNotExist:
                        messages.error(request, f"Please add settings for {school}")
                        return redirect('settingshomepage')

                        pass
                except Exception as exception:
                    messages.error(request, f"{exception}")
                    return redirect('settingshomepage')


            except Exception as exception:
                messages.error(request, exception)
                return redirect('settingshomepage')

            try:
                mobileInstance.save()
                messages.error(request, 'Mobile was updated!')
                return redirect('settingshomepage')
            except:
                messages.error(request, 'Invalid entry')
                return redirect('settingshomepage')

    return redirect('settingshomepage')